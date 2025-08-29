# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io
import openpyxl

class VotingProcessImportWizard(models.TransientModel):
    _name = 'voting.process.import.wizard'
    _description = 'Import voting processes from XLSX'

    file = fields.Binary(string='XLSX File')
    filename = fields.Char(string='File Name')

    def download_template(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _('Processes')
        ws.append(['description', 'date_start', 'date_end', 'tz'])
        ws.append(['Example Election', '2025-09-01 08:00:00', '2025-09-01 18:00:00', 'America/Santiago'])
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        data = base64.b64encode(stream.read())
        self.file = data
        self.filename = _('sample_template.xlsx')
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/?model=%s&id=%s&field=file&download=true&filename=%s' % (self._name, self.id, self.filename),
            'target': 'self',
        }

    def import_processes(self):
        if not self.file:
            raise UserError(_('Debe subir un archivo XLSX.'))
        stream = io.BytesIO(base64.b64decode(self.file))
        wb = openpyxl.load_workbook(stream)
        ws = wb.active
        processes = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            description, date_start, date_end, tz = row
            if not description or not date_start or not date_end or not tz:
                raise UserError(_('Fila %s incompleta.') % idx)
            processes.append({
                'description': description,
                'date_start': date_start,
                'date': date_end,
                'tz': tz,
                'state': 'draft',
            })
        self.env['voting.process'].sudo().create(processes)
        return {'type': 'ir.actions.act_window_close'}
